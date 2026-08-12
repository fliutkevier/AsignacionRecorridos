"""Lectura de la lista de colectores (xlsx con Numero | Colector).

Alimenta la hoja COLECTORES del archivo final y el desplegable de la columna COLECTOR.
"""
from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

import config

log = logging.getLogger(__name__)


def _celda(ws, fila: int, col: int) -> str:
    v = ws.cell(fila, col).value
    return "" if v is None else str(v).strip()


def _buscar_hoja(wb):
    """Por encabezado, no por nombre: el archivo puede venir con otro título de hoja."""
    objetivo = [c.upper() for c in config.COLECTORES_HEADER]
    for ws in wb.worksheets:
        actual = [_celda(ws, 1, i + 1).upper() for i in range(len(objetivo))]
        if actual == objetivo:
            return ws
    return None


def leer(ruta: str | Path) -> list[tuple[str, str]]:
    """Devuelve [(numero, nombre), ...] sin filas vacías ni nombres repetidos.

    El número se conserva como texto para no depender de cómo lo guardó Excel.
    """
    ruta = Path(ruta)
    wb = openpyxl.load_workbook(ruta, data_only=True)
    try:
        ws = _buscar_hoja(wb)
        if ws is None:
            raise ValueError(
                f"'{ruta.name}' no tiene ninguna hoja con el encabezado "
                f"{' | '.join(config.COLECTORES_HEADER)} en la fila 1.")

        vistos: set[str] = set()
        colectores: list[tuple[str, str]] = []
        for f in range(2, ws.max_row + 1):
            nombre = _celda(ws, f, 2)
            if not nombre or nombre in vistos:
                continue                     # fila vacía o nombre repetido
            vistos.add(nombre)
            colectores.append((_celda(ws, f, 1), nombre))
    finally:
        wb.close()

    if not colectores:
        raise ValueError(f"'{ruta.name}' no tiene ningún colector cargado.")

    log.info("Colectores: %d", len(colectores))
    return colectores
