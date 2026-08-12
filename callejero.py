"""Lectura del callejero (xlsx TURNO_NN__dd-mm_WR.xlsx) y armado del índice.

Encabezado real: TURNO | RUTA_EMA | BARRIO | CALLE | ALT_MIN | ALT_MAX | c
  - RUTA_EMA es el RECORRIDO (el lote), no una ruta.
  - Hay filas de subtotal intercaladas ("Total 400") con TURNO vacío: se descartan.
  - La paridad NO viene: se deduce del último dígito de ALT_MIN. Verificado consistente
    con ALT_MAX en 881/881 filas de los turnos 37 y 38.
"""
from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

import config
import texto
from modelos import Tramo

log = logging.getLogger(__name__)

# (barrio, calle) -> {paridad: Tramo}
Indice = dict[tuple[str, str], dict[str, Tramo]]


def _celda(ws, fila: int, col: int) -> str:
    v = ws.cell(fila, col).value
    return "" if v is None else str(v).strip()


def _buscar_hoja(wb):
    """La hoja se busca por su encabezado, nunca por nombre ni por índice."""
    objetivo = [c.upper() for c in config.CALLEJERO_HEADER[:4]]
    for ws in wb.worksheets:
        actual = [_celda(ws, 1, i + 1).upper() for i in range(len(objetivo))]
        if actual == objetivo:
            return ws
    return None


def _paridad(altura: str) -> str:
    """Impar/Par según el último dígito de la altura. '00205' -> 'I'."""
    for ch in reversed(altura):
        if ch.isdigit():
            return "I" if int(ch) % 2 else "P"
    return "P"      # sin dígitos: caso no visto en datos reales


def leer(ruta: str | Path) -> list[Tramo]:
    ruta = Path(ruta)
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=False)
    try:
        ws = _buscar_hoja(wb)
        if ws is None:
            raise ValueError(
                f"'{ruta.name}' no tiene ninguna hoja con el encabezado "
                f"{' | '.join(config.CALLEJERO_HEADER[:4])} en la fila 1.")

        col = config.CALLEJERO_COL
        tramos: list[Tramo] = []
        for f in range(2, ws.max_row + 1):
            turno = _celda(ws, f, col["TURNO"])
            calle = _celda(ws, f, col["CALLE"])
            if not turno or not calle:
                continue                     # fila de subtotal ("Total 400")

            rec = _celda(ws, f, col["RUTA_EMA"])
            if not rec.isdigit():
                raise ValueError(
                    f"'{ruta.name}' fila {f}: RUTA_EMA = {rec!r} no es numérico. "
                    "El recorrido se escribe como número en la salida.")

            alt_min = _celda(ws, f, col["ALT_MIN"])
            medidores = _celda(ws, f, col["c"])
            tramos.append(Tramo(
                turno=turno,
                recorrido=int(rec),
                barrio=_celda(ws, f, col["BARRIO"]),
                calle=calle,
                alt_min=alt_min,
                alt_max=_celda(ws, f, col["ALT_MAX"]),
                paridad=_paridad(alt_min),
                medidores=int(medidores) if medidores.isdigit() else 0,
            ))
    finally:
        wb.close()

    if not tramos:
        raise ValueError(f"'{ruta.name}' no tiene filas de datos.")

    turnos = {config.turno_norm(t.turno) for t in tramos}
    if len(turnos) > 1:
        raise ValueError(
            f"'{ruta.name}' mezcla más de un turno: {', '.join(sorted(turnos))}. "
            "El callejero tiene que ser de un solo turno.")

    log.info("Callejero: turno %s, %d tramos, %d recorridos",
             next(iter(turnos)), len(tramos), len({t.recorrido for t in tramos}))
    return tramos


def indexar(tramos: list[Tramo]) -> Indice:
    """Índice (barrio, calle) -> {paridad: tramo}.

    Cada tramo se registra bajo la variante cruda Y la des-mojibakeada, de barrio y de
    calle. Así no hay que detectar si el texto está corrupto: matchea por cualquiera de
    las dos. Sin esto el match cae de 99,8% a 91,6%.
    """
    idx: Indice = {}
    for t in tramos:
        for b in texto.variantes(t.barrio):
            for c in texto.variantes(t.calle):
                idx.setdefault((b, c), {})[t.paridad] = t
    return idx
