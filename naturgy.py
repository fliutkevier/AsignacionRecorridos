"""Lectura del export del portal, en CSV (camino principal) o xlsx (si lo convirtieron).

Todo se devuelve como str: los ceros a la izquierda de la ruta ('0087') no se pueden perder.

El CSV es UTF-8 CON BOM, separador coma, CRLF, todos los campos entrecomillados. La
codificación va declarada explícita: leerlo con la del sistema rompe 'CERVIÑO' y se
pierden matches en silencio.
"""
from __future__ import annotations

import csv
import logging
from dataclasses import dataclass
from pathlib import Path

import openpyxl

import config

log = logging.getLogger(__name__)


@dataclass(slots=True)
class Tabla:
    """Contenido crudo del export: encabezado + filas, todo texto."""
    encabezado: list[str]
    filas: list[list[str]]


def leer(ruta: str | Path) -> Tabla:
    ruta = Path(ruta)
    tabla = _leer_csv(ruta) if ruta.suffix.lower() == ".csv" else _leer_xlsx(ruta)
    log.info("Naturgy: %d rutas", len(tabla.filas))
    return tabla


# ------------------------------------------------------------------------- CSV
def _leer_csv(ruta: Path) -> Tabla:
    with ruta.open(encoding=config.NATURGY_ENCODING, newline="") as f:
        filas = [fila for fila in csv.reader(f) if fila]

    if not filas:
        raise ValueError(f"'{ruta.name}' está vacío.")

    hdr = [c.strip() for c in filas[0]]
    _validar(hdr, ruta)
    i_ruta = _indice(hdr, "Ruta")

    datos = [_ajustar(f, len(hdr)) for f in filas[1:]
             if len(f) > i_ruta and f[i_ruta].strip()]
    return Tabla(hdr, datos)


# ------------------------------------------------------------------------ XLSX
def _leer_xlsx(ruta: Path) -> Tabla:
    wb = openpyxl.load_workbook(ruta, data_only=True)
    try:
        # Al convertir el CSV, Excel trunca el nombre de hoja a 31 caracteres y suele
        # dejar una 'Hoja1' vacía al lado. Ni por nombre ni por índice: por encabezado.
        ws = _buscar_hoja(wb)
        if ws is None:
            raise ValueError(
                f"'{ruta.name}' no tiene ninguna hoja con el encabezado "
                f"{' | '.join(config.NATURGY_HEADER_MINIMO)} en la fila 1.")

        ancho = ws.max_column
        hdr = [_celda(ws, 1, c) for c in range(1, ancho + 1)]
        _validar(hdr, ruta)
        i_ruta = _indice(hdr, "Ruta")

        datos = []
        for f in range(2, ws.max_row + 1):
            fila = [_celda(ws, f, c) for c in range(1, ancho + 1)]
            if fila[i_ruta]:
                datos.append(fila)
    finally:
        wb.close()

    return Tabla(hdr, datos)


def _buscar_hoja(wb):
    objetivo = [c.upper() for c in config.NATURGY_HEADER_MINIMO]
    for ws in wb.worksheets:
        actual = [_celda(ws, 1, i + 1).upper() for i in range(len(objetivo))]
        if actual == objetivo:
            return ws
    return None


def _celda(ws, fila: int, col: int) -> str:
    v = ws.cell(fila, col).value
    return "" if v is None else str(v).strip()


# ------------------------------------------------------------------------ común
def _validar(hdr: list[str], ruta: Path) -> None:
    bajas = [h.lower() for h in hdr]
    for esperada in config.NATURGY_HEADER_MINIMO:
        if esperada.lower() not in bajas:
            raise ValueError(
                f"'{ruta.name}' no tiene la columna {esperada!r}.\n"
                f"  Encabezado leído: {' | '.join(hdr)}\n"
                "  Si abriste el CSV con doble clic y lo guardaste, es probable que Excel "
                "haya usado ';' como separador y todo haya quedado en una sola columna.")


def indice(hdr: list[str], nombre: str) -> int:
    """Índice 0-based de una columna por nombre, sin distinguir mayúsculas."""
    return _indice(hdr, nombre)


def _indice(hdr: list[str], nombre: str) -> int:
    objetivo = nombre.lower()
    for i, h in enumerate(hdr):
        if h.lower() == objetivo:
            return i
    return -1


def _ajustar(fila: list[str], ancho: int) -> list[str]:
    if len(fila) < ancho:
        return fila + [""] * (ancho - len(fila))
    return fila[:ancho]
