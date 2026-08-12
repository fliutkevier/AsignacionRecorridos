"""Escritura del xlsx de salida: RUTAS, REVISAR y FUERA_NATURGY.

Los archivos de entrada nunca se tocan.
"""
from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config
import texto
from modelos import Estado, Resultado, Resumen, Tramo

log = logging.getLogger(__name__)

_FMT_TEXTO = "@"
_FMT_NUMERO = "0"
_FILL_ENCABEZADO = PatternFill("solid", fgColor=config.COLOR_ENCABEZADO)
_FILL_REVISAR = PatternFill("solid", fgColor=config.COLOR_REVISAR)


def escribir(destino: str | Path, columnas: list[str], pos_insercion: int,
             filas: list[Resultado], fuera: list[Tramo],
             colectores: list[tuple[str, str]] | None = None) -> Resumen:
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)

    encabezado = [*columnas, *config.COLS_NUEVAS]
    i_ruta = len(columnas) + 1                  # columna RUTA, 1-based
    i_colector = encabezado.index(config.COL_COLECTOR) + 1
    # Columnas que hay que emitir como número (índices 1-based dentro del encabezado).
    numericas = [i + 1 for i, c in enumerate(columnas) if c in config.COLS_NUMERICAS]
    # MIN/MAX conservan los ceros a la izquierda del callejero: van como texto.
    texto_cols = [pos_insercion + 1, pos_insercion + 2, i_ruta]

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = config.HOJA_RUTAS
    _volcar(ws, encabezado, filas, pos_insercion, texto_cols, numericas, resaltar=True)

    ws2 = wb.create_sheet(config.HOJA_REVISAR)
    _volcar(ws2, encabezado, [f for f in filas if f.estado == Estado.REVISAR],
            pos_insercion, texto_cols, numericas, resaltar=False)

    ws3 = wb.create_sheet(config.HOJA_FUERA)
    _volcar_fuera(ws3, fuera)

    if colectores:
        ws4 = wb.create_sheet(config.HOJA_COLECTORES)
        _volcar_colectores(ws4, colectores)
        # El desplegable va solo en RUTAS: REVISAR es una copia de diagnóstico y editarla
        # no cambia nada en la hoja principal.
        _agregar_desplegable(ws, i_colector, len(filas), len(colectores))

    wb.save(destino)

    a_revisar = sum(1 for f in filas if f.estado == Estado.REVISAR)
    return Resumen(total=len(filas), asignadas=len(filas) - a_revisar,
                   a_revisar=a_revisar, fuera_naturgy=len(fuera))


def _volcar_colectores(ws, colectores: list[tuple[str, str]]) -> None:
    ws.append(list(config.COLS_COLECTORES))
    for numero, nombre in colectores:
        ws.append([numero, nombre])
    _forzar_texto(ws, 1)          # el número puede tener ceros a la izquierda
    _formatear(ws, len(config.COLS_COLECTORES))


def _agregar_desplegable(ws, col: int, n_filas: int, n_colectores: int) -> None:
    """Validación de lista en la columna COLECTOR, apuntando a la hoja COLECTORES.

    Se referencia el rango y no una lista literal: los 118 nombres suman ~1.900
    caracteres y Excel corta las listas embebidas en 255.
    """
    letra = get_column_letter(col)
    origen = (f"'{config.HOJA_COLECTORES}'!$B$2:$B${n_colectores + 1}")
    dv = DataValidation(type="list", formula1=origen, allow_blank=True, showDropDown=False)
    dv.errorTitle = "Colector inválido"
    dv.error = "Elegí un colector de la lista desplegable."
    dv.promptTitle = "Colector"
    dv.prompt = "Elegí de la lista."
    ws.add_data_validation(dv)
    # Se cubren filas de más para que el desplegable siga funcionando si alguien
    # agrega renglones a mano.
    ultima = n_filas + 1 + config.FILAS_VALIDACION_EXTRA
    dv.add(f"{letra}2:{letra}{ultima}")


def _volcar(ws, encabezado: list[str], filas: list[Resultado], pos_insercion: int,
            texto_cols: list[int], numericas: list[int], resaltar: bool) -> None:
    ws.append(encabezado)
    for f in filas:
        ws.append(f.fila(pos_insercion))
        if resaltar and f.estado == Estado.REVISAR:
            for c in range(1, len(encabezado) + 1):
                ws.cell(ws.max_row, c).fill = _FILL_REVISAR

    # RUTA y MIN/MAX como texto: si Excel los toma como número, '0087' -> 87 y en el
    # turno 43 (rutas de 3 dígitos) el cero perdido ya no se puede reponer.
    for col in texto_cols:
        _forzar_texto(ws, col)
    for col in numericas:
        _forzar_numero(ws, col)
    _formatear(ws, len(encabezado))


def _volcar_fuera(ws, fuera: list[Tramo]) -> None:
    ws.append(list(config.COLS_FUERA))
    for t in fuera:
        ws.append([
            t.recorrido, t.barrio_limpio, t.calle_limpia, t.alt_min, t.alt_max,
            t.paridad, t.medidores,
            "tramo del callejero sin correspondencia en el export de Naturgy" + t.sugerencia,
        ])

    for col in (4, 5):                          # ALT_MIN / ALT_MAX: ceros a la izquierda
        _forzar_texto(ws, col)
    _formatear(ws, len(config.COLS_FUERA))


def _forzar_texto(ws, col: int) -> None:
    letra = get_column_letter(col)
    for fila in range(2, ws.max_row + 1):
        celda = ws[f"{letra}{fila}"]
        celda.number_format = _FMT_TEXTO
        if celda.value is not None:
            celda.value = str(celda.value)


def _forzar_numero(ws, col: int) -> None:
    """Convierte a int lo que el lector trajo como texto (todo el export viene en str).
    Si algún valor no es numérico se deja como está: mejor un texto visible que un 0
    inventado."""
    letra = get_column_letter(col)
    for fila in range(2, ws.max_row + 1):
        celda = ws[f"{letra}{fila}"]
        v = celda.value
        if v is None or isinstance(v, int):
            continue
        s = str(v).strip()
        if s.lstrip("-").isdigit():
            celda.value = int(s)
            celda.number_format = _FMT_NUMERO


def _formatear(ws, columnas: int) -> None:
    for c in range(1, columnas + 1):
        celda = ws.cell(1, c)
        celda.font = Font(name=config.FUENTE, bold=True, color="FFFFFF", size=config.TAM_FUENTE)
        celda.fill = _FILL_ENCABEZADO
        celda.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(columnas)}{ws.max_row}"
        fuente = Font(name=config.FUENTE, size=config.TAM_FUENTE)
        for fila in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=columnas):
            for celda in fila:
                celda.font = fuente

    tope = min(ws.max_row, 300)
    for c in range(1, columnas + 1):
        letra = get_column_letter(c)
        largo = max((len(str(ws[f"{letra}{f}"].value or "")) for f in range(1, tope + 1)),
                    default=config.ANCHO_MIN)
        ws.column_dimensions[letra].width = min(max(largo + 2, config.ANCHO_MIN),
                                                config.ANCHO_MAX)
